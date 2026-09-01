#!/usr/bin/env python3
"""Rebuild data/0AD_civ_strengths.xlsx from data/sheets/*.csv."""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHEETS = [
    "readme",
    "scorecard",
    "scoring_rules",
    "all_units",
    "citizen_infantry",
    "champions",
    "siege_and_crush",
    "ships",
    "houses",
    "buildings",
    "gather_rates",
    "civ_bonuses",
    "civ-gated_techs",
    "all_techs",
    "roster_matrix",
]
TITLES = {
    "readme": "Readme",
    "scorecard": "Scorecard",
    "scoring_rules": "Scoring rules",
    "all_units": "All units",
    "citizen_infantry": "Citizen infantry",
    "champions": "Champions",
    "siege_and_crush": "Siege and crush",
    "ships": "Ships",
    "houses": "Houses",
    "buildings": "Buildings",
    "gather_rates": "Gather rates",
    "civ_bonuses": "Civ bonuses",
    "civ-gated_techs": "Civ-gated techs",
    "all_techs": "All techs",
    "roster_matrix": "Roster matrix",
}


def load_sheet(src, key):
    fp = src / f"{key}.csv"
    p1 = src / f"{key}_part1.csv"
    p2 = src / f"{key}_part2.csv"
    if fp.exists():
        return pd.read_csv(fp)
    if p1.exists() and p2.exists():
        return pd.concat([pd.read_csv(p1), pd.read_csv(p2)], ignore_index=True)
    return None


def main():
    root = Path(__file__).resolve().parents[1]
    src = root / "data" / "sheets"
    out = root / "data" / "0AD_civ_strengths.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for key in SHEETS:
            df = load_sheet(src, key)
            if df is None:
                continue
            name = TITLES[key]
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.book[name[:31]]
            hdr = Font(name="Calibri", color="FFFFFF", bold=True)
            fill = PatternFill("solid", fgColor="1F4E79")
            for cell in ws[1]:
                cell.font = hdr
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for i, col in enumerate(ws.columns, start=1):
                width = 12
                for cell in list(col)[:30]:
                    if cell.value is not None:
                        width = max(width, min(40, len(str(cell.value)) + 2))
                ws.column_dimensions[get_column_letter(i)].width = width
    print("wrote", out)


if __name__ == "__main__":
    main()
